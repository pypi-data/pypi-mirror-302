# -*- coding: utf-8 -*-

# @File    : actual.py
# @Date    : 2023-01-09
# @Author  : Dovelet
"""
因为总会有奇怪的bug，所以打算搞一个快乐的类
"""
import os
import cx_Oracle
import numpy as np
from openpyxl import Workbook
import datetime
import pandas as pd
import netCDF4
from datetime import timedelta
from christmas.commonCode import osprint


class Measured_port:
    def __init__(self, name, user, password, ip, value, sql):

        self.name=name
        self.user=user
        self.password=password
        self.ip=ip
        self.value=value
        self.sql=sql

    @staticmethod
    def mkdir(path):
        # sourcery skip: remove-unnecessary-else, swap-if-else-branches, use-named-expression
        """
        :param path:文件夹路径
        :return: 1-执行成功  0-执行失败
        """
        # 去除首位空格
        path = path.strip()
        # 去除尾部 \ 符号
        path = path.rstrip("\\")
        # 判断路径是否存在
        # 存在     True
        # 不存在   False
        isExists = os.path.exists(path)
        # 判断结果
        if not isExists:
            # 如果不存在则创建目录
            # 创建目录操作函数
            os.makedirs(path)
            return True
        else:
            return False

    @staticmethod
    def hourNums(startTime, endTime):
        """
        :param startTime:开始时间
        :param endTime:截止时间
        :return:两个时间点相差的小时数
        """
        startTime2 = datetime.datetime.strptime(startTime, "%Y-%m-%d %H:%M:%S")
        # 注意，seconds获得的秒只是时间差中的小时、分钟和秒部分的和，并没有包含时间差的天数（既是两个时间点不是同一天，失效）
        total_seconds = (endTime - startTime2).total_seconds()
        # 来获取准确的时间差，并将时间差转换为小时
        hours = total_seconds / 3600
        return int(hours)

    def connector(self, _date):
        # sourcery skip: for-append-to-extend, simplify-generator
        if len(_date) != 8:
            today = datetime.date.today()  # 运行日期，从运行当天开始
            ago = today - timedelta(days=1)

        else:
            today = str(_date)
            today = datetime.datetime.strptime(today, '%Y%m%d')
            ago = today - timedelta(days=1)
            today = today.strftime("%Y-%m-%d")

        # print(today)

        this_month_start = datetime.datetime(ago.year, ago.month, 1)
        m_s = this_month_start.strftime('%Y%m%d')  # 这个月的第一天
        realDIR1 = f'/home/ocean/Oceanmax/Data/real/{self.name}/xlsx'  # 实测数据存放路径
        realDIR2 = f'/home/ocean/Oceanmax/Data/real/{self.name}/nc'
        self.mkdir(realDIR1)
        self.mkdir(realDIR2)
        wb = Workbook()
        ws = wb.active
        ws.title = self.name
        datalist = []
        table_name = ['ID', 'WAVE_DATE', 'HM0', 'H10', 'HMAX', 'HMEAN', 'H3', 'TP', 'TM02', 'TZ', 'DIRTP', 'SPRTP',
                      'MEANDIR', 'UPDATE_DATE', 'LONGITUDE', 'LATITUDE']
        for i in range(len(table_name)):
            ws.cell(row=1, column=i+1, value=table_name[i])
        with cx_Oracle.connect(self.user, self.password, self.ip) as connection:
            cursor = connection.cursor()
            sql=self.sql

            sql3 = 'select * from ' + sql + ' where WAVE_DATE >= to_date(' + m_s + ',\'yyyy-mm-dd\')'
            # print(sql3)
            for result in cursor.execute(sql3):
                datalist.append(result)

        for i in range(len(datalist)):
            for j in range(len(datalist[i])):
                ws.cell(row=i + 2, column=j + 1, value=datalist[i][j])

        needday=str(today)
        wb.save(f'{realDIR1}/{self.name}-{needday}.xlsx')
        # wb.save(realDIR1 + '/djk01-' + str(today) + '.xlsx')

        # 对数据排序 其实数据库那就可以用order排序，但是我就想多学几个库
        stexcel = pd.read_excel(f'{realDIR1}/{self.name}-{needday}.xlsx')
        stexcel.sort_values(by='WAVE_DATE', inplace=True, ascending=True)
        stexcel.to_excel(f'{realDIR1}/{self.name}-{needday}.xlsx')
        # 保存为nc文件

        path = f'{realDIR1}/{self.name}-{needday}.xlsx'
        data = pd.DataFrame(pd.read_excel(path))  # 读取数据,设置None可以生成一个字典，字典中的key值即为sheet名字，此时不用使用DataFram，会报错
        # print(data.index)#获取行的索引名称
        # print(data.columns)#获取列的索引名称

        data=data.drop_duplicates("WAVE_DATE")

        data["WAVE_DATE"] = data["WAVE_DATE"].astype("datetime64[ns]")  # 确保数据格式为日期
        date_range = pd.date_range(start=data["WAVE_DATE"].min(), end=data["WAVE_DATE"].max(),
                                   freq="H")  # freq="D"表示按天，可以按分钟，月，季度，年等
        data = data.set_index("WAVE_DATE").reindex(index=date_range)

        H3 = data[self.value]  # 获取这一列的内容
        TM02 = data['TM02']
        wave_date = date_range  # data['WAVE_DATE']

        hs = []
        for i in range(len(H3)):
            if H3[i] == -9:
                hs.append(np.nan)
            else:
                hs.append(H3[i])

        # print(data.loc[0])#获取行名为0这一行的内容
        startTime = '2000-01-01 00:00:00'

        date = []
        for i in range(len(H3)):
            hour = self.hourNums(startTime, wave_date[i])
            date.append(hour)

        t02 = [TM02[i] for i in range(len(H3))]
        nc_path=f'{realDIR2}/{self.name}-{needday}.nc'
        osprint(f"'{self.name}-{needday}.nc'")
        ncfile = netCDF4.Dataset(nc_path, 'w', format='NETCDF3_64BIT')
        ncfile.createDimension('hour_long', len(H3))

        ncfile.createVariable('H3', np.float64, 'hour_long')
        ncfile.createVariable('time', np.float64, 'hour_long')
        ncfile.createVariable('TM02', np.float64, 'hour_long')

        hs = np.array(hs)
        date = np.array(date)
        t02 = np.array(t02)

        ncfile.variables['H3'][:] = hs[:]
        ncfile.variables['time'][:] = date[:]
        ncfile.variables['TM02'][:] = t02[:]

        ncfile.close()
