"""This module was written as a PoC for SEED data portal.

In current state this SeedExcelStrategy has no sense. But I hope that I find
resources to rewrite it and create a proper base XLSX strategy.

"""

from __future__ import annotations

import logging
from typing import IO, Callable, Iterable, TypedDict, cast

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ckanext.ingest import shared

log = logging.getLogger(__name__)


class XlsxChunk(TypedDict):
    sheet: Worksheet
    document: Workbook
    locator: Callable[[str], Worksheet | None]


class XlsxStrategy(shared.ExtractionStrategy):
    """Extractor data from XLSX files.

    Options[extras]:

        sheets: list[str] - names of processed sheets. All other sheets are not
        extracted but still available via locator in code.

        min_row/max_row/min_col/max_col: int - restrict the scope of parsing

        with_header: bool - parse rows as dict, using the first row for names.
        When unset, every row parsed as a list.

    """

    mimetypes = {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}

    def chunks(
        self,
        source: shared.Storage,
        options: shared.StrategyOptions,
    ) -> Iterable[XlsxChunk]:
        doc = load_workbook(cast(IO[bytes], source), read_only=True, data_only=True)
        sheets = options.get("extras", {}).get("sheets", doc.sheetnames)

        for sheet in doc:
            if sheet.title not in sheets:
                continue

            yield {
                "sheet": sheet,
                "document": doc,
                "locator": lambda name: doc.get(name, None),  # type: ignore
            }

    def extract(
        self,
        source: shared.Storage,
        options: shared.StrategyOptions,
    ) -> Iterable[shared.Record]:
        extras = options.get("extras", {})

        for chunk in self.chunks(source, options):
            rows = chunk["sheet"].iter_rows(
                min_row=extras.get("min_row"),
                max_row=extras.get("max_row"),
                min_col=extras.get("min_col"),
                max_col=extras.get("max_col"),
            )
            header = None
            if extras.get("with_header"):
                header = [c.value for c in next(rows)]

            for row in rows:
                values = [c.value for c in row]
                data = dict(zip(header, values)) if header else values

                yield self.chunk_into_record(
                    {"row": data},
                    options,
                )
