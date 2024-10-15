#!/usr/bin/env python3
# -*- coding: utf-8 -*-
""" A class to import DRF forms"""
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Optional, Union

from openpyxl import Workbook, load_workbook, worksheet
from rich.progress import Progress, track

from regscale.core.app.logz import create_logger
from regscale.core.app.utils.app_utils import create_progress_object
from regscale.integrations.integration.issue import IntegrationIssue
from regscale.models.regscale_models.deviation import Deviation
from regscale.models.regscale_models.issue import Issue
from regscale.models.regscale_models.property import Property

DR_SUBMISSION_DATE = "DR Submission Date"
CVSS_BASE_SCORE = "CVSS Base Score"
ADJUSTED_CVSS_SCORE = "Adjusted CVSS Score"
DR_NUMBER = "DR Number"


class DRF(IntegrationIssue):
    """
    Deviation Request Form class
    """

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.properties = []
        logger = create_logger()
        self.logger = logger
        file_path: Union[str, None] = kwargs.get("file_path", None)
        if not file_path:
            raise ValueError("File path is required")
        self.file_path = Path(file_path)
        self.id_other_identifier_map = self.get_id_map()
        self.module = kwargs.get("module", "securityplans")
        self.module_id = kwargs.get("module_id", 0)

        self.metadata = {}
        self.report = []
        self.errors = []
        self.existing_deviations = Deviation.get_existing_deviations_by_ssp(
            self.module_id,
            issue_ids=[poam.get("id") for poam in self.id_other_identifier_map.values()],
            poam_map=self.id_other_identifier_map,
        )
        self.drf_data = []
        self.missed_drf_data = []
        self.import_drf()
        self.logger.info("Saving DR Identifiers to Issue properties.. ")
        create_properties = [
            prop
            for prop in self.properties
            if prop.value.upper()
            not in {
                val.get("dr_number").upper() for val in self.id_other_identifier_map.values() if val.get("dr_number")
            }
        ]
        update_properties = self.get_update_properties()
        if create_properties:
            Property.batch_create(create_properties)
        if update_properties:
            Property.batch_update(update_properties)
        self.export_report()

    def pull(self) -> tuple[Workbook, list[str]]:
        """
        Pull data from Excel Workbook

        :return: The workbook and the sheets
        :rtype: tuple[Workbook, list[str]]
        """
        workbook = load_workbook(filename=self.file_path, data_only=True, read_only=True)
        sheets = workbook.sheetnames
        pattern = "DR Sheet"

        return workbook, [item for item in sheets if re.search(pattern, item)]

    def gen_metadata(self, ws: worksheet._read_only.ReadOnlyWorksheet) -> dict:
        """
        Generate metadata from the worksheet

        :param worksheet._read_only.ReadOnlyWorksheet ws: The worksheet
        :return: The metadata
        :rtype: dict
        """

        def pull_row_value(row: int, column: int) -> str:
            """
            A method to pull a value from a row and column

            :param int row: The row number
            :param int column: The column number
            :return: The value of the cell
            :rtype: str
            """
            return ws.cell(row=row + 1, column=column + 1).value

        return {
            "CSP Name": pull_row_value(2, 0),
            "System Name": pull_row_value(2, 1),
            "Impact Level": pull_row_value(2, 2),
            DR_SUBMISSION_DATE: pull_row_value(2, 3).strftime("%Y-%m-%d"),
            "CSP Primary POC": pull_row_value(5, 0),
            "Title": pull_row_value(5, 1),
            "Phone": pull_row_value(5, 2),
            "Email": pull_row_value(5, 3),
        }

    def import_drf(self) -> Workbook:
        """
        Import the DRF

        :rtype: Workbook
        :return: The workbook
        """
        workbook, drf_sheets = self.pull()
        with create_progress_object() as progress:
            for sheet in drf_sheets:
                ws = workbook[sheet]
                self.metadata = self.gen_metadata(ws)  # This will probably be important later
                self.parse_sheet_and_generate_deviations(ws, sheet, progress)
            self.save_deviations(progress)
        return workbook

    def parse_sheet_and_generate_deviations(self, ws: worksheet, sheet: str, progress: Progress):
        """
        Parse the sheet and generate deviations

        :param worksheet ws: The worksheet
        :param str sheet: The sheet name
        :param Progress progress: The progress object
        """
        parsing_drfs = progress.add_task(f"[#ef5d23]Parsing '{sheet}' sheet for DRFs...", total=ws.max_row)
        columns = [(cell.value.replace("\n", "")).strip() for cell in ws[9] if cell.value]
        min_row = 0
        for index, row in enumerate(ws.iter_rows(min_row=0, max_row=ws.max_row, values_only=True)):
            if index > min_row:
                if row[0] and isinstance(row[0], str) and row[0].upper().startswith("DR-"):
                    drf = self.gen_drf_from_row(columns=columns, row=row, index=index + min_row, sheet=sheet)
                    if drf:
                        self.drf_data.append(drf)
                    else:
                        self.missed_drf_data.append(row)
                progress.update(parsing_drfs, advance=1)
        self.logger.info("Found %s Deviations ready to create or update", len(self.drf_data))

    def save_deviations(self, progress: Progress) -> None:
        """
        Save the deviations to RegScale

        :param Progress progress: The progress object
        :rtype: None
        """
        saving_drfs = progress.add_task(
            "[#D9F837]Creating or Updating Deviations in RegScale...", total=len(self.drf_data)
        )
        for dev in self.drf_data:
            if dev.extra_data["dr_number"] in {ex.extra_data["dr_number"] for ex in self.existing_deviations}:
                dev.id = [
                    item.id
                    for item in self.existing_deviations
                    if item.extra_data["dr_number"].upper() == dev.extra_data["dr_number"].upper()
                ].pop()
                dev.save()
            else:
                dev.create()
            progress.update(saving_drfs, advance=1)

    def get_property(self, dr_number: str, matching_poam_id: int):
        """
        Get the property and update the list
        """
        prop = Property(
            name=DR_NUMBER,
            key="dr_number",
            value=dr_number,
            parentId=matching_poam_id,
            parentModule="issues",
            label="Deviation Request Number",
            isPublic=True,
        )
        self.properties.append(prop)

    def gen_drf_from_row(self, columns: list[str], row: tuple, index: int, sheet: str) -> Optional[Deviation]:
        """
        Generate a Deviation from a row

        :param list[str] columns: The columns
        :param tuple row: The row
        :param int index: The index
        :param str sheet: The sheet

        :return: The Deviation or None
        :rtype: Optional[Deviation]
        """

        def get_val(index_str: str, default_val: Optional[Any] = None) -> Optional[Any]:
            """
            Get the value from the row

            :param str index_str: The index string
            :param Optional[Any] default_val: The default value to use if nothing is found
            :return: The value or None
            :rtype: Optional[Any]
            """
            index_str = (str(index_str)).strip() if index_str and isinstance(index_str, str) else ""
            try:
                if (dat := row[columns.index(index_str)]) is not None:
                    return str(dat)
            except ValueError:
                self.logger.error("Unable to find column %s in sheet %s", index_str, sheet)
            except TypeError:
                self.logger.error("Type Error: %s, %s", index, sheet)
            return str(default_val) if default_val else None

        # Unique Ident Coalfire
        dr_number = get_val(DR_NUMBER).upper().strip() if get_val(DR_NUMBER) else ""
        poam_id = get_val("POA&M ID").upper().strip() if get_val("POA&M ID") else ""

        matching_poam = self.id_other_identifier_map.get(poam_id)
        if not matching_poam:
            self.report.append({"dr_number": dr_number, "poam_id": poam_id, "status": "Unmatched"})
            return
        matching_poam_id = matching_poam.get("id")

        self.get_property(dr_number=dr_number, matching_poam_id=matching_poam_id)

        deviation_type = Deviation.mapping().get(get_val("Type of DR"))
        justification = get_val("Justification")
        if deviation_type == "Risk Adjustment (RA)" and not justification:
            justification = "Unknown Justification"
        requested_risk_rating = (
            get_val("Requested Risk Rating/Impact") if get_val("Requested Risk Rating/Impact") else "Low"
        )
        self.report.append({"dr_number": dr_number, "poam_id": poam_id, "status": "Matched"})
        if requested_risk_rating and requested_risk_rating.lower() not in ["low", "moderate", "high"]:
            self.logger.error("A valid Requested Risk Rating is required for %s", dr_number)
            self.errors.append(
                {"dr_number": dr_number, "error": f"The Requested Risk Rating {requested_risk_rating} is invalid"}
            )
            return
        if not deviation_type:
            self.logger.error("Unable to find deviation type for %s", dr_number)
            self.errors.append({"dr_number": dr_number, "error": "Deviation Type not found"})
            return None
        if deviation_type == "Risk Adjustment (RA)" and not justification:
            self.logger.error("Justification is required for RA Deviation %s", dr_number)
            self.errors.append({"dr_number": dr_number, "error": f"Justification is required for {deviation_type}"})
            return None
        return Deviation(
            id=0,
            otherIdentifier=poam_id,
            extra_data={"dr_number": dr_number},
            baseScore=(
                float(get_val(CVSS_BASE_SCORE))
                if get_val(CVSS_BASE_SCORE) and str(get_val(CVSS_BASE_SCORE)).isnumeric()
                else None
            ),
            environmentalScore=(
                float(get_val(ADJUSTED_CVSS_SCORE))
                if get_val(ADJUSTED_CVSS_SCORE) and str(get_val(ADJUSTED_CVSS_SCORE)).isnumeric()
                else None
            ),
            parentIssueId=matching_poam_id,
            isPublic=True,
            deviationType=deviation_type,
            requestedImpactRating=requested_risk_rating,
            dateSubmitted=(
                get_val(DR_SUBMISSION_DATE).strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "Z"
                if isinstance(get_val(DR_SUBMISSION_DATE), datetime)
                else None
            ),  # Format must be 2024-08-15T19:00:26.372Z
            rationale=get_val("DR Rationale", ""),
            evidenceDescription=(get_val("Evidence Description") or get_val("Tool-Provided Vulnerability Description")),
            operationalImpacts=(
                get_val("Operational Impact Statement") if get_val("Operational Impact Statement") else "N/A"
            ),
            riskJustification=justification,
            tmpExploitCodeMaturity=None,
            tmpRemediationLevel=get_val("Remediation Level"),
            tmpReportConfidence=None,
            envConfidentiality=get_val("Impact Metrics: Confidentiality"),
            envIntegrity=get_val("Impact Metrics: Integrity"),
            envAvailability=get_val("Impact Metrics: Availability"),
            envAttackVector=get_val("Attack Vector"),
            envAttackComplexity=get_val("Attack Complexity"),
            envPrivilegesRequired=get_val("Privileges Required"),
            envUserInteraction=get_val("User Interaction"),
            envScope=None,
            envModConfidentiality=None,
            envModIntegrity=None,
            envModAvailability=None,
            vulnerabilityId=get_val("Vulnerability Name"),
            envAttackVectorExplanation=get_val("Attack Vector Explanation"),
            envAttackComplexityExplanation=get_val("Attack Complexity Explanation"),
            envPrivilegesRequiredExplanation=get_val("Privileges Required Explanation"),
            envUserInteractionExplanation=get_val("User Interaction Explanation"),
            envConfidentialityExplanation=get_val("Impact Metrics: Confidentiality Explanation"),
            envIntegrityExplanation=get_val("Impact Metrics: Integrity Explanation"),
            envAvailabilityExplanation=get_val("Impact Metrics: Availability Explanation"),
            tmpExploitCodeMaturityExplanation=None,
            tmpRemediationLevelExplanation=get_val("Remediation Level Explanation"),
            tmpReportConfidenceExplanation=None,
            baseSeverity=get_val("Initial Risk Rating"),
            temporalSeverity=None,
            environmentalSeverity=None,
            finalVectorString=None,
            overallRiskReductionExplanation=get_val("List of Risk Reduction"),
            additionalInformation=get_val("List of Evidence Attachments"),
        )

    def get_id_map(self) -> dict:
        """
        Get the ID map

        :return: The ID map
        :rtype: dict
        """
        self.logger.info(f"Fetching all issues for {self.module} #{self.module_id}...")
        id_map = {}
        all_issues = Issue.get_all_by_parent(parent_id=self.module_id, parent_module=self.module)
        self.logger.info("Fetched %s issue(s) from RegScale.", len(all_issues))

        for issue in track(all_issues, description="Building id-otherIdentifier lookup..."):
            if issue.otherIdentifier:
                ident = issue.otherIdentifier.upper()
                id_map[ident] = {"id": issue.id, "dr_number": None, "prop_id": None}

                properties = Property.get_all_by_parent(parent_id=issue.id, parent_module="issues")
                dr_number_property = next((prop for prop in properties if prop.key == "dr_number"), None)
                if dr_number_property:
                    id_map[ident]["dr_number"] = dr_number_property.value.upper()
                    id_map[ident]["prop_id"] = dr_number_property.id
        self.logger.info(
            "Constructed a map of %s issues with POAM id (otherIdentifier) and a nested dictionary of dr_number, issue_key, and property_key",
            len(id_map),
        )
        return id_map

    def export_report(self):
        """Save a Report of missing POAMs to a file"""
        timestamp = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
        report_file = self.file_path.parent / f"DRF-Import-Report_{timestamp}.csv"
        with open(report_file, "w") as file:
            file.write("DR Number,POAM ID,Status,Errors\n")
            for item in self.report:
                item["errors"] = []
                errors = [er for er in self.errors if er["dr_number"] == item["dr_number"]]
                if errors:
                    for error in errors:
                        item["status"] = "Error"
                        item["errors"].append(error)
                if item["status"] in ["Unmatched", "Error"]:
                    file.write(f"{item['dr_number']},{item['poam_id']},{item['status']},{item['errors']}\n")
        self.logger.info("Mismatched POAM Report saved to %s", report_file)

    def get_update_properties(self) -> list[Property]:
        """
        Get the properties to update

        :return: The properties to update
        :rtype: list[Property]
        """
        # Filter properties that have a corresponding 'dr_number' and 'prop_id' in 'id_other_identifier_map'
        props = [
            prop
            for prop in self.properties
            if prop.value
            in {
                val.get("dr_number")
                for val in self.id_other_identifier_map.values()
                if val.get("dr_number") and val.get("prop_id")
            }
        ]

        # Iterate over the filtered properties
        for prop in props:
            # Find the matching value in 'id_other_identifier_map' where 'dr_number' equals the property value
            match = next(
                (val for val in self.id_other_identifier_map.values() if val.get("dr_number") == prop.value), None
            )

            # If a match is found, update the property id
            if match:
                prop.id = match.get("prop_id")

        # Return the list of properties that have an id
        return [prop for prop in props if prop.id]
