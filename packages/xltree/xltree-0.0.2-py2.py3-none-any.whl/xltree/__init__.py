import os
import datetime
import openpyxl as xl

from src.xltree.database import TreeTable
from src.xltree.workbooks import TreeDrawer, TreeEraser


class Config():
    """構成"""


    def __init__(
        self,
        dictionary=None):
        """初期化
        
        Parameters
        ----------
        dictionary : dict
            設定

            列の幅設定。width はだいたい 'ＭＳ Ｐゴシック' サイズ11 の半角英文字の個数
            * `no_width` - A列の幅。no列
            * `row_header_separator_width` - B列の幅。空列
            * `node_width` - 例：C, F, I ...列の幅。ノードの箱の幅
            * `parent_side_edge_width` - 例：D, G, J ...列の幅。エッジの水平線のうち、親ノードの方
            * `child_side_edge_width` - 例：E, H, K ...列の幅。エッジの水平線のうち、子ノードの方

            行の高さ設定。height の単位はポイント。既定値 8。昔のアメリカ人が椅子に座ってディスプレイを見たとき 1/72 インチに見える大きさが 1ポイント らしいが、そんなんワカラン。目視確認してほしい
            * `header_height` - 第１行。ヘッダー
            * `column_header_separator_height` - 第２行。空行
        """

        # 既定のディクショナリー
        self._dictionary = {
            # 列の幅
            'no_width':                         4,
            'row_header_separator_width':       3,
            'node_width':                       20,
            'parent_side_edge_width':           2,
            'child_side_edge_width':            4,

            # 行の高さ
            'header_height':                    13,
            'column_header_separator_height':   13,
        }

        # 上書き
        if dictionary is not None:
            for key, value in dictionary.items():
                self._dictionary[key] = value


    @property
    def dictionary(self):
        return self._dictionary


class WorkbookControl():
    """ワークブック制御"""


    def __init__(self, target, config=Config(), debug_write=False):
        """初期化

        Parameters
        ----------
        target : str
            ワークブック（.xlsx）へのファイルパス
        """
        self._wb_file_path = target
        self._config = config
        self._debug_write = debug_write
        self._wb = None
        self._ws = None


    @property
    def workbook_file_path(self):
        return self._wb_file_path


    def render_sheet(self, target, based_on, debug_write=False):
        """シート描画

        Parameters
        ----------
        target : str
            シート名
        based_on : str
            CSVファイルパス
        debug_write : bool
            デバッグライト
        """

        if self._wb is None:
            self.ready_workbook()

        self.ready_worksheet(target=target)

        # CSV読込
        tree_table = TreeTable.from_csv(file_path=based_on)

        # ツリードロワーを用意、描画（都合上、要らない罫線が付いています）
        tree_drawer = TreeDrawer(tree_table=tree_table, ws=self._ws, config=self._config, debug_write=debug_write)
        tree_drawer.render()


        # 要らない罫線を消す
        # DEBUG_TIPS: このコードを不活性にして、必要な線は全部描かれていることを確認してください
        if True:
            tree_eraser = TreeEraser(tree_table=tree_table, ws=self._ws, debug_write=debug_write)
            tree_eraser.render()
        else:
            if self._debug_write:
                print(f"[{datetime.datetime.now()}] eraser disabled")


    def remove_worksheet(self, target):
        """存在すれば、指定のワークシートの削除。存在しなければ無視

        Parameters
        ----------
        target : str
            シート名
        """

        if self.exists_sheet(target=target):
            if self._debug_write:
                print(f"[{datetime.datetime.now()}] remove `{target}` sheet...")

            self._wb.remove(self._wb[target])


    def save_workbook(self):
        """ワークブックの保存"""

        if self._debug_write:
            print(f"[{datetime.datetime.now()}] save `{self._wb_file_path}` file...")

        # ワークブックの保存            
        self._wb.save(self._wb_file_path)


    def ready_workbook(self):
        """ワークブックを準備する"""

        # 既存のファイルがあるなら読込
        if os.path.isfile(self._wb_file_path):
            if self._debug_write:
                print(f"[{datetime.datetime.now()}] `{self._wb_file_path}` file exists, read.")

            self._wb = xl.load_workbook(filename=self._wb_file_path)
        
        # ワークブックを新規生成
        else:
            if self._debug_write:
                print(f"[{datetime.datetime.now()}] `{self._wb_file_path}` file not exists, create.")

            self._wb = xl.Workbook()


    def exists_sheet(self, target):
        """シートの存在確認
        
        Parameters
        ----------
        target : str
            シート名
        """
        return target in self._wb.sheetnames


    def ready_worksheet(self, target):
        """ワークシートを準備する
        
        Parameters
        ----------
        target : str
            シート名
        """

        # シートを作成
        if not self.exists_sheet(target):
            if self._debug_write:
                print(f"[{datetime.datetime.now()}] create `{target}` sheet...")

            self._wb.create_sheet(target)


        self._ws = self._wb[target]
